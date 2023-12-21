#importações
import openpyxl

#criando planilha inicial
planilha = openpyxl.Workbook()

#Criando uma página dentro da minha planilha
# planilha.create_sheet('Troque pelo nome da sua página') 

# #O código acima ajuda a criar páginas dentro da planilha chamada na primeira linha deste código

planilha.create_sheet('Fuzis')
planilha.create_sheet('Livros')
planilha.create_sheet('Dinheiro')
planilha.create_sheet('Frutas')

#selecionando uma página dentro da planilha
page_frutas = planilha ['Frutas']
#page_fuzis = planilha ['Fuzis']
#page_Dinheiro = planilha ['Dinheiro']
#page_livros = planilha ['Dinheiro']


#adicionar dados em uma página através de linhas e colunas.
#Para adicionar uma coluna devemos duplicar este código abaixo e alterar os valores
page_frutas.append(['Fruta','Compra','Unidades','Preço','Reponsável','Saída','Chegada'])
page_frutas.append(['Banana','Ceasa-RJ','1000','5,00','Fulano','05:00','08:00'])
# Salvar a planilha
planilha.save('Planilha com todas as compras.xlsx')

#O código abaixo mostra no terminal o nome das páginas dentro da planilha
#print(planilha.sheetnames)