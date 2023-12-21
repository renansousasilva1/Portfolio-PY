#importação das dependências de python
import openpyxl

#importar/abrir a planilha para editar com python
planilha = openpyxl.load_workbook('Planilha com todas as compras.xlsx')

#Selecionar uma página dentro da planilha
pagina_fruta = planilha['Frutas']

#Editar os dados de cada linha - #Dentro do parantêses estabelecemos o intervalo das linhas
for rows in pagina_fruta.iter_rows(min_row=1, max_row=2):
    for cell in rows:
        if cell.value == 'Banana':
            cell.value = 'Valor para inserir no lugar'
        

#Para salvar a versão editada
planilha.save('Planilha com todas as compras V2.xlsx')

#Para visualizar as alterações
for rows in pagina_fruta.iter_rows(min_row=1, max_row=5):
     print(f'{rows[0].value}, {rows[1].value}, {rows[2].value}, {rows[3].value}, {rows[4].value}, {rows[5].value}, {rows[6].value}')


        
#Para imprimir/vis em ordem, um ao lado do outro
#for rows in pagina_fruta.iter_rows(min_row=1, max_row=2):
 #       print(f'{rows[0].value}, {rows[1].value}, {rows[2].value}, {rows[3].value}, {rows[4].value}, {rows[5].value}, {rows[6].value}')