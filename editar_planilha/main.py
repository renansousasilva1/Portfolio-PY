#importação das dependências de python
import openpyxl

#importar/abrir a planilha para editar com python
planilha = openpyxl.load_workbook('Planilha com todas as compras.xlsx')

#Selecionar uma página dentro da planilha
pagina_fuzil = planilha['Fuzis']

#Imprimir os dados de cada linha - #Dentro do parantêses estabelecemos o intervalo das linhas
for rows in pagina_fuzil.iter_rows(min_row=1, max_row=2):
    for cell in rows:
        print(cell.value)
        
#Para imprimir em ordem, um ao lado do outro
for rows in pagina_fuzil.iter_rows(min_row=1, max_row=2):
        print(f'{rows[0].value}, {rows[1].value}, {rows[2].value}, {rows[3].value}, {rows[4].value}, {rows[5].value}, {rows[6].value}')